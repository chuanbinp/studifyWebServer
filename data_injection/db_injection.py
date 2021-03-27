#Injection of Questions data into SSAD.Question Collection

import requests
from openpyxl import load_workbook


#Retrieve data from excel sheet
wb_obj = load_workbook(filename = 'db_question.xlsx')
sheet_obj = wb_obj.active
sheet_dims = (sheet_obj.max_row, sheet_obj.max_column)

INDEX = {
    "QUESTION": 1,
    "CATEGORY": 2,
    "DIFFICULTY": 3,
    "OPTION1": 4,
    "OPTION2": 5,
    "OPTION3": 6,
    "OPTION4": 7,
    "ANSWER": 8
}

#Error handling
if(len(INDEX)!=sheet_dims[1]):
    print("Mismatch in number of fields in DB and Excel.")
    print("DB: ", len(INDEX), "; Excel: ", sheet_dims[1])
    print("Please check again! Exiting script...")

#Proceed to inject
else:
    #For each row, process it and inject into db
    for row in range(2,sheet_dims[0]+1):
        curr_row = {}
        curr_row["question"] = sheet_obj.cell(row,INDEX["QUESTION"]).value
        curr_row["category"] = sheet_obj.cell(row,INDEX["CATEGORY"]).value
        curr_row["difficulty"] = sheet_obj.cell(row,INDEX["DIFFICULTY"]).value
        curr_row["options"] = [
            sheet_obj.cell(row,INDEX["OPTION1"]).value,
            sheet_obj.cell(row,INDEX["OPTION2"]).value,
            sheet_obj.cell(row,INDEX["OPTION3"]).value,
            sheet_obj.cell(row,INDEX["OPTION4"]).value,
        ]
        curr_row["answer"] = sheet_obj.cell(row,INDEX["ANSWER"]).value

        #Post request
        response = requests.post('http://localhost:5000/questions/', data = curr_row)
        if(response.status_code==201):
            print("Succesfully injected row ", row-1,"/", sheet_obj.max_row-1, ".")
        else:
            print("Injection of row ", row-1, " failed. Exiting script...")

    print("Successfully injected all data! Exiting script...")
        
   